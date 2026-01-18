#!/usr/bin/env python3
"""Create Excel dashboard from session analysis data."""

import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Installing openpyxl...")
    import subprocess

    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

# Read the CSV data
csv_path = Path("session_analysis.csv")
sessions = []
with open(csv_path, "r") as f:
    reader = csv.DictReader(f)
    for row in reader:
        sessions.append(row)

# Create workbook
wb = Workbook()
if wb.active:
    wb.remove(wb.active)  # Remove default sheet

# ===== SHEET 1: Summary Dashboard =====
ws_summary = wb.create_sheet("Dashboard", 0)

# Title
ws_summary["A1"] = "Amplifier Problem-Solving Approaches Dashboard"
ws_summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
ws_summary["A1"].fill = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
ws_summary.merge_cells("A1:F1")
ws_summary.row_dimensions[1].height = 30

# Metadata
ws_summary["A3"] = "Analysis Date:"
ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
ws_summary["A4"] = "Total Sessions:"
ws_summary["B4"] = len(sessions)
ws_summary["A5"] = "Date Range:"
dates = [s["Created"] for s in sessions if s["Created"]]
if dates:
    ws_summary["B5"] = f"{min(dates)[:10]} to {max(dates)[:10]}"

# Count approaches
approach_counter = Counter()
for session in sessions:
    approaches = session.get("All Approaches", "")
    if approaches:
        for approach in approaches.split(", "):
            approach_counter[approach.strip()] += 1

# ===== SHEET 2: Approach Frequency Data =====
ws_freq = wb.create_sheet("Approach Frequency")

# Headers
ws_freq["A1"] = "Problem-Solving Approach"
ws_freq["B1"] = "Count"
ws_freq["C1"] = "Percentage"
for cell in ["A1", "B1", "C1"]:
    ws_freq[cell].font = Font(bold=True)
    ws_freq[cell].fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

# Data
row = 2
total_sessions = len(sessions)
for approach, count in approach_counter.most_common():
    ws_freq[f"A{row}"] = approach
    ws_freq[f"B{row}"] = count
    ws_freq[f"C{row}"] = f"{(count / total_sessions) * 100:.1f}%"
    row += 1

# Adjust column widths
ws_freq.column_dimensions["A"].width = 35
ws_freq.column_dimensions["B"].width = 12
ws_freq.column_dimensions["C"].width = 12

# Add bar chart to summary
chart = BarChart()
chart.title = "Problem-Solving Approach Frequency"
chart.x_axis.title = "Approach"
chart.y_axis.title = "Number of Sessions"
data = Reference(ws_freq, min_col=2, min_row=1, max_row=row - 1)
cats = Reference(ws_freq, min_col=1, min_row=2, max_row=row - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 12
chart.width = 20
ws_summary.add_chart(chart, "A8")

# ===== SHEET 3: Primary Approach Distribution =====
ws_primary = wb.create_sheet("Primary Approach")

# Count primary approaches
primary_counter = Counter()
for session in sessions:
    primary = session.get("Primary Approach", "Unknown")
    if primary:
        primary_counter[primary] += 1

# Headers
ws_primary["A1"] = "Primary Approach"
ws_primary["B1"] = "Count"
for cell in ["A1", "B1"]:
    ws_primary[cell].font = Font(bold=True)
    ws_primary[cell].fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

# Data
row = 2
for approach, count in primary_counter.most_common():
    ws_primary[f"A{row}"] = approach
    ws_primary[f"B{row}"] = count
    row += 1

ws_primary.column_dimensions["A"].width = 35
ws_primary.column_dimensions["B"].width = 12

# Add pie chart to summary
pie = PieChart()
pie.title = "Primary Approach Distribution"
data = Reference(ws_primary, min_col=2, min_row=1, max_row=row - 1)
cats = Reference(ws_primary, min_col=1, min_row=2, max_row=row - 1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
pie.height = 12
pie.width = 20
ws_summary.add_chart(pie, "K8")

# ===== SHEET 4: Time-based Analysis =====
ws_time = wb.create_sheet("Timeline")

# Group by date
sessions_by_date = defaultdict(list)
for session in sessions:
    if session.get("Created"):
        date = session["Created"][:10]
        sessions_by_date[date].append(session)

# Headers
ws_time["A1"] = "Date"
ws_time["B1"] = "Total Sessions"
ws_time["C1"] = "Exploratory"
ws_time["D1"] = "Error Recovery"
ws_time["E1"] = "Validation"
ws_time["F1"] = "Direct Implementation"
for cell in ["A1", "B1", "C1", "D1", "E1", "F1"]:
    ws_time[cell].font = Font(bold=True)
    ws_time[cell].fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

# Data
row = 2
for date in sorted(sessions_by_date.keys()):
    day_sessions = sessions_by_date[date]
    ws_time[f"A{row}"] = date
    ws_time[f"B{row}"] = len(day_sessions)

    # Count specific approaches
    exploratory = sum(
        1 for s in day_sessions if "Exploratory" in s.get("All Approaches", "")
    )
    error_recovery = sum(
        1 for s in day_sessions if "Error Recovery" in s.get("All Approaches", "")
    )
    validation = sum(
        1 for s in day_sessions if "Validation" in s.get("All Approaches", "")
    )
    direct = sum(
        1
        for s in day_sessions
        if "Direct Implementation" in s.get("All Approaches", "")
    )

    ws_time[f"C{row}"] = exploratory
    ws_time[f"D{row}"] = error_recovery
    ws_time[f"E{row}"] = validation
    ws_time[f"F{row}"] = direct
    row += 1

for col in ["A", "B", "C", "D", "E", "F"]:
    ws_time.column_dimensions[col].width = 15

# ===== SHEET 5: Success Patterns =====
ws_success = wb.create_sheet("Success Patterns")

ws_success["A1"] = "Pattern Analysis"
ws_success["A1"].font = Font(size=14, bold=True)
ws_success.merge_cells("A1:C1")

row = 3
ws_success[f"A{row}"] = "Metric"
ws_success[f"B{row}"] = "Value"
ws_success[f"C{row}"] = "Notes"
for cell in [f"A{row}", f"B{row}", f"C{row}"]:
    ws_success[cell].font = Font(bold=True)
    ws_success[cell].fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

row += 1
# Sessions with file modifications
file_mod = sum(
    1 for s in sessions if "Files Modified" in s.get("Success Indicators", "")
)
ws_success[f"A{row}"] = "Sessions with File Modifications"
ws_success[f"B{row}"] = file_mod
ws_success[f"C{row}"] = f"{(file_mod / total_sessions) * 100:.1f}%"

row += 1
# Sessions with validation
validated = sum(1 for s in sessions if "Validated" in s.get("Success Indicators", ""))
ws_success[f"A{row}"] = "Sessions with Validation"
ws_success[f"B{row}"] = validated
ws_success[f"C{row}"] = f"{(validated / total_sessions) * 100:.1f}%"

row += 1
# Sessions with good error recovery
error_rec = sum(
    1 for s in sessions if "Good Error Recovery" in s.get("Success Indicators", "")
)
ws_success[f"A{row}"] = "Sessions with Good Error Recovery"
ws_success[f"B{row}"] = error_rec
ws_success[f"C{row}"] = f"{(error_rec / total_sessions) * 100:.1f}%"

row += 1
# Substantial work sessions
substantial = sum(
    1 for s in sessions if "Substantial Work" in s.get("Success Indicators", "")
)
ws_success[f"A{row}"] = "Substantial Work Sessions"
ws_success[f"B{row}"] = substantial
ws_success[f"C{row}"] = f"{(substantial / total_sessions) * 100:.1f}%"

row += 2
ws_success[f"A{row}"] = "Average Metrics"
ws_success[f"A{row}"].font = Font(bold=True)
ws_success.merge_cells(f"A{row}:C{row}")

row += 1
# Average turns
avg_turns = sum(int(s.get("Turn Count", 0) or 0) for s in sessions) / total_sessions
ws_success[f"A{row}"] = "Average Turns per Session"
ws_success[f"B{row}"] = f"{avg_turns:.1f}"

row += 1
# Average messages
avg_messages = (
    sum(int(s.get("Message Count", 0) or 0) for s in sessions) / total_sessions
)
ws_success[f"A{row}"] = "Average Messages per Session"
ws_success[f"B{row}"] = f"{avg_messages:.1f}"

for col in ["A", "B", "C"]:
    ws_success.column_dimensions[col].width = 35

# ===== SHEET 6: Raw Data =====
ws_raw = wb.create_sheet("Raw Data")

# Read CSV and write to sheet
with open(csv_path, "r") as f:
    reader = csv.reader(f)
    for r_idx, row in enumerate(reader, 1):
        for c_idx, value in enumerate(row, 1):
            ws_raw.cell(row=r_idx, column=c_idx, value=value)

# Format headers
for col in range(1, len(sessions[0]) + 1):
    cell = ws_raw.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Freeze top row
ws_raw.freeze_panes = "A2"

# Save
output_path = (
    Path.home() / "Downloads" / "amplifier-sessions-problem-solving-dashboard.xlsx"
)
wb.save(output_path)
print(f"✅ Dashboard created: {output_path}")
print(f"📊 Analyzed {total_sessions} sessions")
print(f"📈 {len(approach_counter)} unique approaches identified")
